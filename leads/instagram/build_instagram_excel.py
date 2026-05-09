"""
Fusiona los 3 outputs (paso 1 metadata, paso 2 perfiles, paso 3 contact info)
y genera el Excel final con marca de fuente por dato.

Si algún paso no se ejecutó, simplemente se omite (los datos quedan "pendiente").
"""
import json, os
from urllib.parse import urlparse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "leads/instagram"
PROFILES_FILE = f"{BASE}/profiles_enriched.json"
STEP1_META    = f"{BASE}/step1_unique_profiles.json"
STEP3_OUT     = f"{BASE}/step3_output.json"          # opcional
OUTPUT_XLSX   = f"{BASE}/leads_instagram_50.xlsx"

def load_json(path, default):
    if not os.path.exists(path):
        return default
    with open(path, encoding="utf-8") as f:
        return json.load(f)

profiles = load_json(PROFILES_FILE, [])
step1    = load_json(STEP1_META, [])
step3    = load_json(STEP3_OUT, [])

# Mapa hashtag/origen por username (paso 1)
hashtag_by_user = {p["username"]: p.get("fromHashtag", "") for p in step1}

# Mapa contact-info por dominio (paso 3)
def domain_of(url):
    if not url: return ""
    try:
        return urlparse(url if url.startswith("http") else "https://"+url).netloc.lower().replace("www.","")
    except Exception:
        return ""

contact_by_domain = {}
for entry in step3:
    d = domain_of(entry.get("url") or entry.get("startUrl") or entry.get("domain", ""))
    if not d: continue
    contact_by_domain[d] = {
        "emails":     entry.get("emails", []) or [],
        "phones":     entry.get("phones", []) or [],
        "instagrams": entry.get("instagrams", []) or [],
        "facebooks":  entry.get("facebooks", []) or [],
    }

# Construir filas finales
rows = []
for p in profiles:
    web = p.get("website", "")
    d   = domain_of(web)
    contact = contact_by_domain.get(d, {})

    # Email: primero el del IG profile, luego el del contact-info
    email     = p.get("businessEmail") or (contact.get("emails", [""])[:1] or [""])[0]
    email_src = ("instagram_profile_scraper" if p.get("businessEmail")
                 else ("contact_info_scraper" if email else "pendiente"))

    # Phone: primero el del IG profile, luego el del contact-info
    phone     = p.get("businessPhone") or (contact.get("phones", [""])[:1] or [""])[0]
    phone_src = ("instagram_profile_scraper" if p.get("businessPhone")
                 else ("contact_info_scraper" if phone else "pendiente"))

    rows.append({
        "username":      p.get("username", ""),
        "nombre":        p.get("fullName", "") or "pendiente",
        "bio":           p.get("bio", "") or "pendiente",
        "seguidores":    p.get("followers", 0),
        "website":       web or "pendiente",
        "correo":        email or "pendiente",
        "telefono":      phone or "pendiente",
        "instagramLink": p.get("instagramLink", ""),
        "hashtag":       hashtag_by_user.get(p.get("username",""), ""),
        "src_nombre":    p.get("_src_fullName", "pendiente"),
        "src_bio":       p.get("_src_bio", "pendiente"),
        "src_website":   p.get("_src_website", "pendiente"),
        "src_correo":    email_src,
        "src_telefono":  phone_src,
    })

# Si hay menos de 50, completar con pendientes (placeholder)
# (en la práctica suele ser 30-50 dependiendo del scraping)

# Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Leads Instagram"

headers = [
    "#", "Username", "Nombre", "Bio", "Seguidores",
    "Website", "Correo", "Teléfono / WhatsApp", "Link Instagram", "Hashtag origen",
    "Fuente Nombre", "Fuente Bio", "Fuente Website", "Fuente Correo", "Fuente Teléfono"
]

header_fill  = PatternFill("solid", fgColor="6B2DB5")
header_font  = Font(bold=True, color="FFFFFF", size=10)
alt_fill     = PatternFill("solid", fgColor="F4ECFB")
white_fill   = PatternFill("solid", fgColor="FFFFFF")
pending_fill = PatternFill("solid", fgColor="FFF2CC")
center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
left         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Side(style="thin", color="C9B6E0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.row_dimensions[1].height = 32
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border

for i, r in enumerate(rows, 1):
    row = i + 1
    fill = alt_fill if i % 2 == 0 else white_fill
    ws.row_dimensions[row].height = 28
    vals = [
        i,
        r["username"],
        r["nombre"],
        (r["bio"][:200] + "…") if len(r["bio"]) > 200 else r["bio"],
        r["seguidores"],
        r["website"],
        r["correo"],
        r["telefono"],
        r["instagramLink"],
        r["hashtag"],
        r["src_nombre"],
        r["src_bio"],
        r["src_website"],
        r["src_correo"],
        r["src_telefono"],
    ]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = border
        c.alignment = left if col not in (1,5) else center
        is_pending = (str(val) == "pendiente")
        if is_pending:
            c.fill = pending_fill
            c.font = Font(color="7F6000", italic=True, size=9)
        else:
            c.fill = fill
            c.font = Font(size=9)

# Anchos
widths = [4, 22, 24, 50, 11, 30, 28, 22, 32, 18, 22, 22, 22, 22, 22]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w
ws.freeze_panes = "A2"

# Hoja leyenda
lg = wb.create_sheet("Leyenda")
lg["A1"] = "Fuentes de datos por columna"; lg["A1"].font = Font(bold=True, size=12)
legend = [
    ("instagram_hashtag_scraper", "Paso 1 — usernames descubiertos por hashtag"),
    ("instagram_profile_scraper", "Paso 2 — bio, seguidores, website, email/phone de cuenta business"),
    ("contact_info_scraper",      "Paso 3 — email/teléfono extraído del website del perfil"),
    ("pendiente",                 "Sin dato — NO se inventa nada"),
]
for r_idx, (a,b) in enumerate(legend, 3):
    lg.cell(row=r_idx, column=1, value=a).font = Font(bold=True, size=9)
    lg.cell(row=r_idx, column=2, value=b).font = Font(size=9)
lg.column_dimensions["A"].width = 32
lg.column_dimensions["B"].width = 70

wb.save(OUTPUT_XLSX)
print(f"✓ Excel generado: {OUTPUT_XLSX}")
print(f"✓ Leads totales:  {len(rows)}")
print(f"  - con correo:    {sum(1 for r in rows if r['correo']!='pendiente')}")
print(f"  - con teléfono:  {sum(1 for r in rows if r['telefono']!='pendiente')}")
print(f"  - con website:   {sum(1 for r in rows if r['website']!='pendiente')}")
