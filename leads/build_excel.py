import json, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

raw = json.load(open("/home/user/Marco/leads/raw_apify.json", encoding="utf-8"))

# Filter only CDMX entries and deduplicate by title+street
seen = set()
leads = []
for r in raw:
    state = r.get("state", "")
    city  = r.get("city", "")
    if "México" not in city and "CDMX" not in city and "CDMX" not in state:
        continue
    key = (r["title"].lower().strip(), r.get("street","").lower().strip())
    if key in seen:
        continue
    seen.add(key)
    leads.append(r)

# Score: phone present (+2), rating weight, reviews log
import math
def score(r):
    s = 0
    if r.get("phone"): s += 2
    s += r.get("totalScore", 0)
    s += math.log1p(r.get("reviewsCount", 0)) * 0.3
    return s

leads.sort(key=score, reverse=True)
leads = leads[:50]

# Build workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Leads CDMX"

headers = [
    "#", "Nombre del Negocio", "Categoría", "Dirección", "Alcaldía / Ciudad",
    "Teléfono", "Correo", "Sitio Web", "Instagram", "Facebook",
    "WhatsApp", "Persona de Contacto", "Rating", "Reseñas",
    "Fuente Teléfono", "Fuente Correo", "Fuente Redes",
    "Estado Enriquecimiento", "Link Google Maps"
]

# Styles
header_fill   = PatternFill("solid", fgColor="1F4E79")
header_font   = Font(bold=True, color="FFFFFF", size=10)
alt_fill      = PatternFill("solid", fgColor="EBF3FB")
white_fill    = PatternFill("solid", fgColor="FFFFFF")
pending_fill  = PatternFill("solid", fgColor="FFF2CC")
center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Side(style="thin", color="B0C4DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header row
ws.row_dimensions[1].height = 32
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = center_align
    c.border = border

# Data rows
for i, r in enumerate(leads, 1):
    row = i + 1
    fill = alt_fill if i % 2 == 0 else white_fill
    ws.row_dimensions[row].height = 22

    phone    = r.get("phone", "") or ""
    tel_src  = "Google Maps" if phone else ""
    needs    = []
    if not phone:          needs.append("teléfono")
    needs.append("correo"); needs.append("redes")
    estado = "Pendiente enriquecimiento: " + ", ".join(needs) if needs else "Completo"

    vals = [
        i,
        r.get("title",""),
        r.get("categoryName",""),
        r.get("street",""),
        r.get("city",""),
        phone,
        "pendiente",
        "pendiente",
        "pendiente",
        "pendiente",
        "pendiente",
        "pendiente",
        r.get("totalScore",""),
        r.get("reviewsCount",""),
        tel_src if phone else "pendiente",
        "pendiente",
        "pendiente",
        estado,
        r.get("url",""),
    ]

    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = border
        c.alignment = left_align if col > 2 else center_align
        if str(val) == "pendiente":
            c.fill = pending_fill
            c.font = Font(color="7F6000", italic=True, size=9)
        elif col == 18:  # estado enriquecimiento
            c.fill = pending_fill
            c.font = Font(color="7F6000", size=9)
        else:
            c.fill = fill
            c.font = Font(size=9)

# Column widths
col_widths = [4, 38, 22, 32, 28, 18, 22, 22, 18, 18, 18, 22, 7, 8, 14, 14, 14, 38, 18]
for col, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

# Freeze header
ws.freeze_panes = "A2"

# Legend sheet
lg = wb.create_sheet("Leyenda")
lg["A1"] = "Leyenda de fuentes y colores"
lg["A1"].font = Font(bold=True, size=12)
legend_rows = [
    ("Fondo amarillo", "Dato pendiente de enriquecimiento — NO inventado"),
    ("Fondo blanco/azul claro", "Dato obtenido del scraper original (Google Maps)"),
    ("Fuente = Google Maps", "Dato proveniente del actor compass/crawler-google-places"),
    ("Fuente = pendiente", "Se buscará en siguiente paso de enriquecimiento (Instagram / Facebook / Email Finder)"),
    ("Estado Enriquecimiento", "Indica qué campos faltan para completar el lead"),
]
for r_idx, (col_a, col_b) in enumerate(legend_rows, 3):
    lg.cell(row=r_idx, column=1, value=col_a).font = Font(bold=True, size=9)
    lg.cell(row=r_idx, column=2, value=col_b).font  = Font(size=9)
lg.column_dimensions["A"].width = 28
lg.column_dimensions["B"].width = 65

out = "/home/user/Marco/leads/leads_CDMX_50.xlsx"
wb.save(out)
print(f"Guardado: {out}")
print(f"Total leads: {len(leads)}")

# Console preview (first 10)
print("\n{'#':<3} {'Negocio':<38} {'Categoría':<22} {'Teléfono':<18} {'Rating':<6} {'Reseñas'}")
print("-"*100)
for i, r in enumerate(leads[:10], 1):
    phone = r.get("phone","") or "pendiente"
    print(f"{i:<3} {r['title'][:37]:<38} {r['categoryName'][:21]:<22} {phone:<18} {r.get('totalScore',''):<6} {r.get('reviewsCount','')}")
print(f"\n... y {len(leads)-10} leads más en el archivo Excel.")
